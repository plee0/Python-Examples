### Imports
import datetime
import sys
####
import bs4
from bs4 import BeautifulSoup
####
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
####
import openpyxl
from openpyxl.styles import PatternFill
###
import os
###

# Function to accept metric quality and metric quantity, and fill as required
def metricFill(flag, quantity, metric):
    metricRow1 = 18 #Past Due Work Instructions
    metricRow2 = 39 #Number WI in ENG Review
    metricRow3 = 60 #Number WI Aged 5 Days
    colDateofMonth = 17 #Column "Q"
    first = True
    
    # Set fill color based on whether metric surpasses criteria
    if flag == True:
        color = 'FF0000' #Red
    else:
        color = '00B050' #Green
    fillColor = PatternFill(fill_type='solid', fgColor= color)

    # Each metric uses different fill criteria because the scales are different
    if metric == 1:
        step = int(quantity / 5)
        metricRow = metricRow1
    elif metric == 2:
        metricRow = metricRow2
        if quantity >= 15:
            step = 13 + int( (quantity - 15) / 5 )
        else:
            # A zero metric still results in one fill of "green"
            step = quantity + 1
    elif metric == 3:
        metricRow = metricRow3
        if quantity >= 15:
            step = 13 + int( (quantity - 15) / 5 )
        else:
            # A zero metric still results in one fill of "green"
            step = quantity + 1

    # Fill cells descending (starting at top then filling downwards)
    for row in worksheet.iter_rows(min_row=metricRow - step, max_row = metricRow - 1, min_col = colDateofMonth+todayDate, max_col= colDateofMonth+todayDate):
        for cell in row:
            cell.fill = fillColor
            # For metric 1 only, also type the number into the box
            if metric == 1 and first == True:
                cell.value = quantity
                first = False


now = datetime.datetime.now()
print('Today\'s Date/Time:')
print(now.strftime("%Y-%m-%d %H:%M:%S"))

#Declare Array to store Metric Data
IMetric = []
Dates = []

#Open SpaceWeb with Microsoft Edge Browser
print('Opening Microsoft Edge')
browser = webdriver.Edge()
browser.get('https://spaceweb/')
wait = WebDriverWait(browser, 20)

#User fillable password and username
user_username = sys.argv[1]
user_password = sys.argv[2]

#Automatically login to SpaceWeb
print('Logging in using your credentials')
username = browser.find_element(By.NAME, 'aw_username')
password = browser.find_element(By.NAME, 'aw_password')
username.send_keys(user_username)
password.send_keys(user_password)
password.send_keys(Keys.RETURN)

#Navigate into Planning Tab - Past Due & Enable Draft Checkbox
browser.get('https://spaceweb/planning/travelers/list.cfm?frmFormat=Past%20Due&frmTravelerStatus=Draft,Defined,Revising,Engineering,Quality,Pending&frmDateFilter=PastDue&frmSortOrder=tblTraveler.NeedDate,defSort&frmSortDir=ASC&frmLayout=General')
#PE = browser.find_element(By.ID, 'lblResultCount')
PE = wait.until(EC.element_to_be_clickable((By.ID,'lblResultCount')))
stat = PE.text.split()[0]
#Add number of past due items into Metric Data Array
IMetric.append(stat)
print('There are ', stat, ' items past due')

#Navigate into Planning Tab - Pending ENG Review
#Get the number of items Pending ENG Review
browser.get('https://spaceweb/planning/travelers/list.cfm?frmFormat=Pending%20Engineering%20Review&frmTravelerStatus=Engineering&frmSortOrder=tblTraveler.NeedDate,defSort&frmSortDir=DESC&frmLayout=General')
PE = wait.until(EC.element_to_be_clickable((By.ID,'lblResultCount')))
stat = PE.text.split()[0]
#Add number of items pending engineering into Metric Data Array
IMetric.append(stat)
print('There are ', stat, ' item(s) pending Engineering Review')

#If there are any items pending, navigate to table in current window and extract work instructions
#from table
if int(stat) > 0:
    browser.get('https://spaceweb/planning/travelers/list.cfm?frmView=Begin&frmFormat=Pending Engineering Review&frmLayout=General&frmSource=Main&frmFlagSQL=0&frmTopFocus=1&frmSortOrder=tblTraveler.NeedDate,tblTraveler.TravelerCode,TravelerStatusSort,tblTraveler.TravelerRevision&frmSortDir=DESC&frmLimitResults=1&frmStartRow=1&frmDivisionID=8&frmSiteID=All&frmTravelerStatus=Engineering&frmTravelerPriority=All&frmAuthorID=All&frmQualityReviewID=All&frmEngReviewID=All&frmJobID=All&frmManagerID=All&frmCreateBy=All&frmTypeID=All&frmPrintStatus=All&frmSchedStatus=All&frmDateFilter=None&frmDateThreshold=30&frmDateFilterFrom=01/01/2025&frmDateFilterTo=01/31/2025&frmHasUndefined=0&frmNoOutputParts=0&frmNoNeedDate=0&frmFlagFavorite=0&frmSearchString=')
    html = browser.page_source
    soup = BeautifulSoup(html,'html.parser')
    tableResults = browser.find_element(By.ID, 'tblResults')
    rows = tableResults.find_elements(By.TAG_NAME, "tr")
    # Parse through the SUBMITTED DATE COLUMN
    for x in range(2, len(rows), 4):
        col = rows[x].find_element(By.TAG_NAME, 'span')
        # Not sure why but I experienced a bug here where this item is not indexable or something... it only happened once - and that final third item actually disappeared after an update. Not sure if it's related to that.
        textDate = col.get_attribute('title')
        dateSubmitted = textDate.split(',')[0]
        #Split the date/time and days aging by comma delimiter, then keep only the number
        try:
            daysAging = int(textDate.split(',')[1].split()[0])
        except:
            #Items submitted TODAY do not have an aged date, and therefore are 0 days aginged
            daysAging = 0

        Dates.append(daysAging)
        #print(textDate)

    overDue = sum(1 for i in Dates if i >= 5)
    print('There are ', overDue, ' work instructions aged 5 days or older.')
    #Add number of engineering items past 5 days into Metric Data Array, if any
    IMetric.append(overDue)
else:
    #If there are no items pending Engineering, then the metric must be 0
    IMetric.append('0')
    
### This next section of code is responsible for editing the excel file ###
#Path to DCA Excel File


username = os.getlogin()
# THIS FILE NEEDS TO GET UPDATED EVERY MONTH WHEN THE EXCEL FILE CHANGES
ExcelFile = 'Engineering Tier 2 DCA Jan 2025.xlsx'

#Test File - Make sure that this filepath is correct prior to executing python script
DCAPath = 'C:\\Users\\' + username + '\\Desktop\\' + ExcelFile
print('Updating I-Metric in excel file located at: ', DCAPath)
DCAWorkbook = openpyxl.load_workbook(DCAPath)
worksheet = DCAWorkbook['I-Metric']

todayDate = now.day

#Metric 1, # of Past Due Work Instructions
flag = False
if int(IMetric[0]) >= 15:
    flag = True
metricFill(flag, int(IMetric[0]), 1)
    
#Metric 2, # of WI's in Eng Review
flag = False
if int(IMetric[1]) > 5:
    flag = True
metricFill(flag, int(IMetric[1]), 2)

#Metric 3, # of WI's in Eng Review Aged Past 5 Days
flag = False
if len(IMetric) == 3:
    if int(IMetric[2]) > 2:
        # Trigger red color only if third metric exists (aged 5 days), and exceeds 2 WI's
        flag = True
metricFill(flag, int(IMetric[2]), 3)

DCAWorkbook.save(DCAPath)
print ('I-Metric Automation Complete')
